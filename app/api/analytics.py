from io import BytesIO

from fastapi import APIRouter, Request, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font

from app.database.dependencies import get_db
from app.services.analytics_service import AnalyticsService
from app.services.material_research_service import MaterialResearchService

router = APIRouter(tags=["Analytics"])

templates = Jinja2Templates(directory="app/templates")


# =====================================================
# JSON API — vendor scorecards
# =====================================================

@router.get("/analytics/vendor-scorecards")
def vendor_scorecards(db: Session = Depends(get_db)):
    return AnalyticsService.all_vendor_scorecards(db)


@router.get("/analytics/vendor-scorecard/{vendor_id}")
def vendor_scorecard(vendor_id: int, db: Session = Depends(get_db)):
    try:
        return AnalyticsService.vendor_scorecard(db, vendor_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# =====================================================
# JSON API — price intelligence
# =====================================================

@router.get("/analytics/price-intelligence")
def price_intelligence(db: Session = Depends(get_db)):
    return AnalyticsService.price_intelligence_overview(db)


@router.get("/analytics/price-history")
def price_history(material_name: str, db: Session = Depends(get_db)):
    return AnalyticsService.price_history(db, material_name)


@router.get("/analytics/evaluate-quote")
def evaluate_quote(material_name: str, rate: float, db: Session = Depends(get_db)):
    return AnalyticsService.evaluate_quote_price(db, material_name, rate)


# =====================================================
# JSON API — reports
# =====================================================

@router.get("/analytics/reports")
def reports(db: Session = Depends(get_db)):
    return AnalyticsService.reports_summary(db)


# =====================================================
# JSON API — material research (LLM)
# =====================================================

@router.get("/analytics/material-research")
def material_research(material_name: str):
    return MaterialResearchService.research(material_name)


# =====================================================
# Excel export — spend reports
# =====================================================

@router.get("/analytics/reports/export")
def export_reports(db: Session = Depends(get_db)):
    data = AnalyticsService.reports_summary(db)

    wb = Workbook()

    def _sheet(title, headers, rows):
        ws = wb.create_sheet(title=title)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.value = h
            c.font = Font(bold=True)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).value = val

    # Drop the default empty sheet
    wb.remove(wb.active)

    _sheet(
        "Spend by Project",
        ["Project", "PO Count", "Total Spend"],
        [[r["project_name"], r["po_count"], r["total_spend"]] for r in data["spend_by_project"]]
    )
    _sheet(
        "Spend by Vendor",
        ["Vendor", "PO Count", "Total Spend"],
        [[r["vendor_name"], r["po_count"], r["total_spend"]] for r in data["spend_by_vendor"]]
    )
    _sheet(
        "Spend by Category",
        ["Category", "Item Count", "Total Spend"],
        [[r["category"], r["item_count"], r["total_spend"]] for r in data["spend_by_category"]]
    )
    _sheet(
        "Negotiation Savings",
        ["RFQ ID", "Vendor ID", "Round", "Original", "Agreed", "Saving"],
        [[r["rfq_id"], r["vendor_id"], r["round_number"], r["original_price"],
          r["agreed_price"], r["saving"]] for r in data["negotiation_savings"]["detail"]]
    )
    _sheet(
        "RFQ Turnaround",
        ["RFQ", "Days to Award"],
        [[r["rfq_number"], r["days"]] for r in data["rfq_turnaround"]["detail"]]
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=procurement_reports.xlsx"}
    )


# =====================================================
# HTML Dashboards
# =====================================================

@router.get("/dashboard/analytics", response_class=HTMLResponse)
def analytics_page(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(
        request=request,
        name="analytics_dashboard.html",
        context={
            "request": request,
            "reports": AnalyticsService.reports_summary(db),
            "price_intelligence": AnalyticsService.price_intelligence_overview(db, limit=25)
        }
    )


@router.get("/dashboard/vendor-scorecards", response_class=HTMLResponse)
def vendor_scorecards_page(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(
        request=request,
        name="vendor_scorecards.html",
        context={
            "request": request,
            "scorecards": AnalyticsService.all_vendor_scorecards(db)
        }
    )
